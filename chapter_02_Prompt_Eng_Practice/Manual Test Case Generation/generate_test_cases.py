#!/usr/bin/env python
"""
Generate comprehensive test cases for Watchguard ID requirements in XLSX format.
Following RICE-POT framework and Anti-Hallucination Rules.

Requirements:
1. Rename 'Loyalty ID' to 'Profile ID' and enable multiple Profile IDs per Watchguard ID
2. Replace Shopper ID with 7-Character Alphanumeric Watchguard ID
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

def create_test_cases_workbook():
    wb = Workbook()
    
    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Create sheets for different test categories
    create_ui_test_sheet(wb, "REQ1_ProfileID_UI")
    create_api_test_sheet(wb, "REQ1_ProfileID_API")
    create_ui_test_sheet(wb, "REQ2_WatchguardID_UI")
    create_api_test_sheet(wb, "REQ2_WatchguardID_API")
    create_integration_sheet(wb, "Integration_Tests")
    
    # Save the workbook
    output_path = r"d:\AI 3x Blueprint\Practice\chapter_02_Prompt_Eng_Practice\Manual Test Case Generation\output\Watchguard_TestCases.xlsx"
    wb.save(output_path)
    print(f"✓ Test cases generated successfully: {output_path}")
    return output_path

def setup_styles():
    """Define reusable styles"""
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    priority_high = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    priority_medium = PatternFill(start_color="FFD93D", end_color="FFD93D", fill_type="solid")
    priority_low = PatternFill(start_color="A8E6CF", end_color="A8E6CF", fill_type="solid")
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    center_align = Alignment(horizontal='center', vertical='top', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    return {
        'header_fill': header_fill,
        'header_font': header_font,
        'priority_high': priority_high,
        'priority_medium': priority_medium,
        'priority_low': priority_low,
        'border': border,
        'center_align': center_align,
        'left_align': left_align
    }

def create_ui_test_sheet(wb, sheet_name):
    """Create UI test cases sheet"""
    ws = wb.create_sheet(sheet_name)
    styles = setup_styles()
    
    # Define columns
    columns = [
        "TID", "Scenario", "Test Data", "Test Case Description", 
        "Pre-Condition", "Test Steps", "Expected Result", 
        "Actual Result", "Status", "Priority", "Is Automated", "Requirement"
    ]
    
    # Add header
    for col_num, col_title in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = col_title
        cell.font = styles['header_font']
        cell.fill = styles['header_fill']
        cell.alignment = styles['center_align']
        cell.border = styles['border']
    
    # Set column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 35
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 40
    ws.column_dimensions['G'].width = 35
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 12
    ws.column_dimensions['J'].width = 10
    ws.column_dimensions['K'].width = 12
    ws.column_dimensions['L'].width = 15
    
    # Generate test cases based on sheet name
    test_cases = []
    if "REQ1" in sheet_name:
        test_cases = get_req1_ui_tests()
    elif "REQ2" in sheet_name:
        test_cases = get_req2_ui_tests()
    
    # Add test cases
    for row_num, test_case in enumerate(test_cases, 2):
        for col_num, value in enumerate(test_case.values(), 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.alignment = styles['left_align']
            cell.border = styles['border']
            
            # Apply priority color
            if col_num == 10:  # Priority column
                if value == "High":
                    cell.fill = styles['priority_high']
                elif value == "Medium":
                    cell.fill = styles['priority_medium']
                else:
                    cell.fill = styles['priority_low']
    
    ws.freeze_panes = 'A2'

def create_api_test_sheet(wb, sheet_name):
    """Create API test cases sheet"""
    ws = wb.create_sheet(sheet_name)
    styles = setup_styles()
    
    # Define columns for API testing
    columns = [
        "TID", "Endpoint", "HTTP Method", "Request Body", "Request Headers",
        "Expected Status", "Expected Response", "Test Data", 
        "Actual Status", "Actual Response", "Status", "Priority", "Requirement"
    ]
    
    # Add header
    for col_num, col_title in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = col_title
        cell.font = styles['header_font']
        cell.fill = styles['header_fill']
        cell.alignment = styles['center_align']
        cell.border = styles['border']
    
    # Set column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 35
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 35
    ws.column_dimensions['H'].width = 20
    ws.column_dimensions['I'].width = 12
    ws.column_dimensions['J'].width = 25
    ws.column_dimensions['K'].width = 12
    ws.column_dimensions['L'].width = 10
    ws.column_dimensions['M'].width = 15
    
    # Generate test cases
    test_cases = []
    if "REQ1" in sheet_name:
        test_cases = get_req1_api_tests()
    elif "REQ2" in sheet_name:
        test_cases = get_req2_api_tests()
    
    # Add test cases
    for row_num, test_case in enumerate(test_cases, 2):
        for col_num, value in enumerate(test_case.values(), 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.alignment = styles['left_align']
            cell.border = styles['border']
            
            # Apply priority color
            if col_num == 12:  # Priority column
                if value == "High":
                    cell.fill = styles['priority_high']
                elif value == "Medium":
                    cell.fill = styles['priority_medium']
                else:
                    cell.fill = styles['priority_low']
    
    ws.freeze_panes = 'A2'

def create_integration_sheet(wb, sheet_name):
    """Create integration and smoke tests sheet"""
    ws = wb.create_sheet(sheet_name)
    styles = setup_styles()
    
    columns = [
        "TID", "Test Type", "Scenario", "Pre-Condition", 
        "Test Steps", "Expected Result", "Priority", "Status", "Requirement"
    ]
    
    # Add header
    for col_num, col_title in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = col_title
        cell.font = styles['header_font']
        cell.fill = styles['header_fill']
        cell.alignment = styles['center_align']
        cell.border = styles['border']
    
    # Set column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 40
    ws.column_dimensions['F'].width = 35
    ws.column_dimensions['G'].width = 10
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 20
    
    test_cases = get_integration_tests()
    
    # Add test cases
    for row_num, test_case in enumerate(test_cases, 2):
        for col_num, value in enumerate(test_case.values(), 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.alignment = styles['left_align']
            cell.border = styles['border']
            
            # Apply priority color
            if col_num == 7:  # Priority column
                if value == "High":
                    cell.fill = styles['priority_high']
                elif value == "Medium":
                    cell.fill = styles['priority_medium']
                else:
                    cell.fill = styles['priority_low']
    
    ws.freeze_panes = 'A2'

def get_req1_ui_tests():
    """UI test cases for Requirement 1: Profile ID Linking"""
    return [
        {
            "TID": "REQ1-UI-001",
            "Scenario": "Display Profile ID in UI (Positive)",
            "Test Data": "Existing Loyalty ID record",
            "Test Case Description": "Verify that all occurrences of 'Loyalty ID' label are replaced with 'Profile ID' in UI",
            "Pre-Condition": "User is logged in; Database migrated with Profile ID naming",
            "Test Steps": "1. Navigate to customer profile page\n2. Inspect UI labels and fields\n3. Verify 'Profile ID' label is displayed",
            "Expected Result": "All UI labels show 'Profile ID' instead of 'Loyalty ID'. No 'Loyalty ID' references visible (except in legacy data notes if applicable)",
            "Actual Result": "",
            "Status": "",
            "Priority": "High",
            "Is Automated": "Yes",
            "Requirement": "REQ1-AC1"
        },
        {
            "TID": "REQ1-UI-002",
            "Scenario": "Link Multiple Profile IDs to Single Watchguard ID",
            "Test Data": "Watchguard ID: WAT-ABC123; Multiple Profile IDs: PID-001, PID-002",
            "Test Case Description": "Verify UI allows linking multiple Profile IDs to one Watchguard ID",
            "Pre-Condition": "User has admin privileges; Watchguard ID exists; Multiple Profile ID records exist",
            "Test Steps": "1. Navigate to Watchguard ID management page\n2. Select a Watchguard ID\n3. Click 'Add Profile ID' button\n4. Select first Profile ID (PID-001)\n5. Click 'Add Profile ID' again\n6. Select second Profile ID (PID-002)\n7. Click 'Save'",
            "Expected Result": "Both Profile IDs are linked to the Watchguard ID. Confirmation message displayed. Linked Profile IDs visible in the UI list",
            "Actual Result": "",
            "Status": "",
            "Priority": "High",
            "Is Automated": "Yes",
            "Requirement": "REQ1-AC2"
        },
        {
            "TID": "REQ1-UI-003",
            "Scenario": "Display Multiple Linked Profile IDs",
            "Test Data": "Watchguard ID with 3 linked Profile IDs",
            "Test Case Description": "Verify UI displays all linked Profile IDs for a given Watchguard ID",
            "Pre-Condition": "Watchguard ID has 3 Profile IDs linked; User is on details page",
            "Test Steps": "1. Open Watchguard ID details\n2. Scroll to 'Linked Profile IDs' section\n3. Count and verify all Profile IDs are listed",
            "Expected Result": "All 3 Profile IDs are displayed in a clear, organized list. Each Profile ID shows creation date and status",
            "Actual Result": "",
            "Status": "",
            "Priority": "High",
            "Is Automated": "Yes",
            "Requirement": "REQ1-AC2"
        },
        {
            "TID": "REQ1-UI-004",
            "Scenario": "Remove Profile ID Link (Negative)",
            "Test Data": "Watchguard ID with 2 linked Profile IDs",
            "Test Case Description": "Verify ability to unlink a Profile ID from Watchguard ID",
            "Pre-Condition": "Watchguard ID has multiple linked Profile IDs",
            "Test Steps": "1. Open Watchguard ID details\n2. Locate one Profile ID in the linked list\n3. Click 'Remove' or 'Unlink' action\n4. Confirm removal in dialog",
            "Expected Result": "Profile ID is unlinked successfully. Confirmation displayed. Removed Profile ID no longer appears in linked list",
            "Actual Result": "",
            "Status": "",
            "Priority": "Medium",
            "Is Automated": "Yes",
            "Requirement": "REQ1-AC2"
        },
        {
            "TID": "REQ1-UI-005",
            "Scenario": "Verify Data Integrity After Migration",
            "Test Data": "Legacy Loyalty ID records migrated to Profile ID",
            "Test Case Description": "Verify existing records remain intact after Profile ID migration",
            "Pre-Condition": "Database migration completed; Existing loyalty records present",
            "Test Steps": "1. Navigate to customer records\n2. Select a migrated record\n3. Verify all data fields are intact\n4. Check linked Watchguard ID is unchanged",
            "Expected Result": "All historical data is preserved. No data loss. Watchguard ID association unchanged. All dates and transaction history intact",
            "Actual Result": "",
            "Status": "",
            "Priority": "High",
            "Is Automated": "Yes",
            "Requirement": "REQ1-AC3"
        },
        {
            "TID": "REQ1-UI-006",
            "Scenario": "Invalid Profile ID Link Attempt (Negative)",
            "Test Data": "Invalid/Non-existent Profile ID: INVALID-9999",
            "Test Case Description": "Verify system rejects linking of non-existent Profile IDs",
            "Pre-Condition": "User on Profile ID linking page",
            "Test Steps": "1. Attempt to link a non-existent Profile ID\n2. Submit the form",
            "Expected Result": "Error message displayed: 'Profile ID does not exist'. Link operation rejected. No invalid record created",
            "Actual Result": "",
            "Status": "",
            "Priority": "High",
            "Is Automated": "Yes",
            "Requirement": "REQ1-Validation"
        },
        {
            "TID": "REQ1-UI-007",
            "Scenario": "Duplicate Profile ID Link Prevention (Negative)",
            "Test Data": "Profile ID: PID-001 already linked to Watchguard ID",
            "Test Case Description": "Verify system prevents linking the same Profile ID twice",
            "Pre-Condition": "Profile ID PID-001 is already linked to Watchguard WAT-ABC",
            "Test Steps": "1. Attempt to link PID-001 again to same Watchguard ID\n2. Submit form",
            "Expected Result": "Error message: 'This Profile ID is already linked to this Watchguard ID'. Duplicate link prevented",
            "Actual Result": "",
            "Status": "",
            "Priority": "High",
            "Is Automated": "Yes",
            "Requirement": "REQ1-Validation"
        },
        {
            "TID": "REQ1-UI-008",
            "Scenario": "Verify API and UI Display Consistency",
            "Test Data": "Profile ID linked via API",
            "Test Case Description": "Verify UI reflects changes made via API (and vice versa)",
            "Pre-Condition": "API has linked new Profile ID to Watchguard ID",
            "Test Steps": "1. Link Profile ID via API call\n2. Refresh UI in browser\n3. Verify UI shows newly linked Profile ID",
            "Expected Result": "UI immediately displays the API-linked Profile ID without page refresh (or after refresh). No data sync delay > 5 seconds",
            "Actual Result": "",
            "Status": "",
            "Priority": "Medium",
            "Is Automated": "Yes",
            "Requirement": "REQ1-AC4"
        },
        {
            "TID": "REQ1-UI-009",
            "Scenario": "Search by Profile ID",
            "Test Data": "Profile ID: PID-ABC-12345",
            "Test Case Description": "Verify user can search customers by Profile ID",
            "Pre-Condition": "Search functionality available; Profile ID records in database",
            "Test Steps": "1. Navigate to search page\n2. Enter Profile ID in search field\n3. Click 'Search'\n4. Verify results",
            "Expected Result": "Search returns all records with matching Profile ID. Results displayed clearly with relevant details",
            "Actual Result": "",
            "Status": "",
            "Priority": "Medium",
            "Is Automated": "Yes",
            "Requirement": "REQ1-Enhancement"
        },
        {
            "TID": "REQ1-UI-010",
            "Scenario": "Bulk Link Profile IDs (Performance)",
            "Test Data": "10 Profile IDs to link to single Watchguard ID",
            "Test Case Description": "Verify system can handle bulk linking of multiple Profile IDs",
            "Pre-Condition": "User has bulk link permission; 10 Profile IDs available",
            "Test Steps": "1. Navigate to bulk link page\n2. Upload CSV with 10 Profile IDs\n3. Select Watchguard ID\n4. Submit bulk link request\n5. Monitor completion status",
            "Expected Result": "All 10 Profile IDs linked successfully within 2 seconds. Success notification displayed. Progress indicator shown",
            "Actual Result": "",
            "Status": "",
            "Priority": "Medium",
            "Is Automated": "No",
            "Requirement": "REQ1-Performance"
        }
    ]

def get_req1_api_tests():
    """API test cases for Requirement 1: Profile ID Linking"""
    return [
        {
            "TID": "REQ1-API-001",
            "Endpoint": "/api/v1/profileids",
            "HTTP Method": "GET",
            "Request Body": "N/A",
            "Request Headers": "Authorization: Bearer <token>",
            "Expected Status": "200",
            "Expected Response": "List of Profile IDs with properties: id, name, status, linkedWatchguardId, createdDate",
            "Test Data": "Valid authentication token",
            "Actual Status": "",
            "Actual Response": "",
            "Status": "",
            "Priority": "High",
            "Requirement": "REQ1-API"
        },
        {
            "TID": "REQ1-API-002",
            "Endpoint": "/api/v1/watchguards/{watchguardId}/profileids",
            "HTTP Method": "POST",
            "Request Body": "{\"profileId\": \"PID-001\"}",
            "Request Headers": "Authorization: Bearer <token>, Content-Type: application/json",
            "Expected Status": "201",
            "Expected Response": "{\"success\": true, \"message\": \"Profile ID linked successfully\", \"data\": {watchguardId, profileId, linkedDate}}",
            "Test Data": "Valid Watchguard ID, Valid Profile ID",
            "Actual Status": "",
            "Actual Response": "",
            "Status": "",
            "Priority": "High",
            "Requirement": "REQ1-API-Link"
        },
        {
            "TID": "REQ1-API-003",
            "Endpoint": "/api/v1/watchguards/{watchguardId}/profileids",
            "HTTP Method": "GET",
            "Request Body": "N/A",
            "Request Headers": "Authorization: Bearer <token>",
            "Expected Status": "200",
            "Expected Response": "{\"data\": [{profileId details}], \"count\": X}",
            "Test Data": "Valid Watchguard ID with multiple linked Profile IDs",
            "Actual Status": "",
            "Actual Response": "",
            "Status": "",
            "Priority": "High",
            "Requirement": "REQ1-API-Retrieve"
        },
        {
            "TID": "REQ1-API-004",
            "Endpoint": "/api/v1/watchguards/{watchguardId}/profileids/{profileId}",
            "HTTP Method": "DELETE",
            "Request Body": "N/A",
            "Request Headers": "Authorization: Bearer <token>",
            "Expected Status": "200",
            "Expected Response": "{\"success\": true, \"message\": \"Profile ID unlinked successfully\"}",
            "Test Data": "Valid Watchguard ID, Valid linked Profile ID",
            "Actual Status": "",
            "Actual Response": "",
            "Status": "",
            "Priority": "High",
            "Requirement": "REQ1-API-Unlink"
        },
        {
            "TID": "REQ1-API-005",
            "Endpoint": "/api/v1/watchguards/{watchguardId}/profileids",
            "HTTP Method": "POST",
            "Request Body": "{\"profileId\": \"INVALID-9999\"}",
            "Request Headers": "Authorization: Bearer <token>, Content-Type: application/json",
            "Expected Status": "400",
            "Expected Response": "{\"error\": \"Profile ID does not exist\"}",
            "Test Data": "Valid Watchguard ID, Invalid Profile ID",
            "Actual Status": "",
            "Actual Response": "",
            "Status": "",
            "Priority": "High",
            "Requirement": "REQ1-API-Validation"
        },
        {
            "TID": "REQ1-API-006",
            "Endpoint": "/api/v1/watchguards/{watchguardId}/profileids",
            "HTTP Method": "POST",
            "Request Body": "{\"profileId\": \"PID-001\"}",
            "Request Headers": "Authorization: Bearer <token>, Content-Type: application/json",
            "Expected Status": "409",
            "Expected Response": "{\"error\": \"Profile ID already linked to this Watchguard ID\"}",
            "Test Data": "Profile ID already linked to same Watchguard ID",
            "Actual Status": "",
            "Actual Response": "",
            "Status": "",
            "Priority": "High",
            "Requirement": "REQ1-API-Conflict"
        },
        {
            "TID": "REQ1-API-007",
            "Endpoint": "/api/v1/watchguards/INVALID-ID/profileids",
            "HTTP Method": "POST",
            "Request Body": "{\"profileId\": \"PID-001\"}",
            "Request Headers": "Authorization: Bearer <token>, Content-Type: application/json",
            "Expected Status": "404",
            "Expected Response": "{\"error\": \"Watchguard ID not found\"}",
            "Test Data": "Invalid Watchguard ID",
            "Actual Status": "",
            "Actual Response": "",
            "Status": "",
            "Priority": "High",
            "Requirement": "REQ1-API-NotFound"
        },
        {
            "TID": "REQ1-API-008",
            "Endpoint": "/api/v1/watchguards/{watchguardId}/profileids",
            "HTTP Method": "POST",
            "Request Body": "{\"profileId\": \"PID-001\"}",
            "Request Headers": "Content-Type: application/json",
            "Expected Status": "401",
            "Expected Response": "{\"error\": \"Unauthorized\"}",
            "Test Data": "Missing authorization token",
            "Actual Status": "",
            "Actual Response": "",
            "Status": "",
            "Priority": "High",
            "Requirement": "REQ1-API-Security"
        },
        {
            "TID": "REQ1-API-009",
            "Endpoint": "/api/v1/watchguards/{watchguardId}/profileids/batch",
            "HTTP Method": "POST",
            "Request Body": "{\"profileIds\": [\"PID-001\", \"PID-002\", \"PID-003\"]}",
            "Request Headers": "Authorization: Bearer <token>, Content-Type: application/json",
            "Expected Status": "201",
            "Expected Response": "{\"success\": true, \"linked\": 3, \"failed\": 0}",
            "Test Data": "Valid Watchguard ID, Multiple valid Profile IDs",
            "Actual Status": "",
            "Actual Response": "",
            "Status": "",
            "Priority": "Medium",
            "Requirement": "REQ1-API-Batch"
        },
        {
            "TID": "REQ1-API-010",
            "Endpoint": "/api/v1/migration/loyalty-to-profileid/validate",
            "HTTP Method": "POST",
            "Request Body": "{\"migrationBatchId\": \"BATCH-001\"}",
            "Request Headers": "Authorization: Bearer <token>, Content-Type: application/json",
            "Expected Status": "200",
            "Expected Response": "{\"valid\": true, \"recordsProcessed\": X, \"recordsIntact\": X, \"dataLossDetected\": false}",
            "Test Data": "Valid batch ID after migration",
            "Actual Status": "",
            "Actual Response": "",
            "Status": "",
            "Priority": "High",
            "Requirement": "REQ1-API-Migration-Validation"
        }
    ]

def get_req2_ui_tests():
    """UI test cases for Requirement 2: Watchguard ID"""
    return [
        {
            "TID": "REQ2-UI-001",
            "Scenario": "Display Watchguard ID (7-char alphanumeric)",
            "Test Data": "Sample Watchguard ID: WG7A2B9",
            "Test Case Description": "Verify Watchguard ID (7-character alphanumeric) is displayed in UI",
            "Pre-Condition": "User logged in; Watchguard ID generated and stored",
            "Test Steps": "1. Navigate to customer/order detail page\n2. Locate ID field\n3. Verify displayed ID is 7-character alphanumeric",
            "Expected Result": "Watchguard ID displayed as 7-character alphanumeric (e.g., WG7A2B9). No Shopper ID visible as primary identifier",
            "Actual Result": "",
            "Status": "",
            "Priority": "High",
            "Is Automated": "Yes",
            "Requirement": "REQ2-AC1"
        },
        {
            "TID": "REQ2-UI-002",
            "Scenario": "Verify Uniqueness of Generated Watchguard ID",
            "Test Data": "Multiple newly created records",
            "Test Case Description": "Verify each newly created record has a unique Watchguard ID",
            "Pre-Condition": "Create multiple new customer records",
            "Test Steps": "1. Create first customer record\n2. Note the generated Watchguard ID\n3. Create second customer record\n4. Note the generated Watchguard ID\n5. Compare IDs - verify they are different",
            "Expected Result": "Each record has unique 7-character Watchguard ID. No duplicate IDs assigned",
            "Actual Result": "",
            "Status": "",
            "Priority": "High",
            "Is Automated": "Yes",
            "Requirement": "REQ2-AC3"
        },
        {
            "TID": "REQ2-UI-003",
            "Scenario": "Display Original Shopper ID in Legacy Reference",
            "Test Data": "Migrated record with original Shopper ID: SHP-000123456",
            "Test Case Description": "Verify original Shopper ID is accessible as reference (not primary identifier)",
            "Pre-Condition": "Migration completed; Records show Watchguard ID as primary",
            "Test Steps": "1. Open customer record\n2. Look for 'Legacy Reference' or 'Original ID' section\n3. Verify original Shopper ID is displayed",
            "Expected Result": "Original Shopper ID visible in read-only 'Legacy Reference' section. Clearly marked as historical/reference data",
            "Actual Result": "",
            "Status": "",
            "Priority": "Medium",
            "Is Automated": "Yes",
            "Requirement": "REQ2-AC2"
        },
        {
            "TID": "REQ2-UI-004",
            "Scenario": "Search/Filter by Watchguard ID",
            "Test Data": "Watchguard ID: WG7A2B9",
            "Test Case Description": "Verify user can search and filter records by Watchguard ID",
            "Pre-Condition": "Search functionality available",
            "Test Steps": "1. Navigate to search page\n2. Enter Watchguard ID in search field\n3. Click 'Search'\n4. Verify results",
            "Expected Result": "Correct record returned with matching Watchguard ID. Quick and accurate search results",
            "Actual Result": "",
            "Status": "",
            "Priority": "High",
            "Is Automated": "Yes",
            "Requirement": "REQ2-Enhancement"
        },
        {
            "TID": "REQ2-UI-005",
            "Scenario": "Watchguard ID Format Validation (Negative)",
            "Test Data": "Invalid Watchguard ID: WG7A2 (only 5 chars)",
            "Test Case Description": "Verify system validates Watchguard ID format",
            "Pre-Condition": "User attempting to manually enter or modify Watchguard ID",
            "Test Steps": "1. Attempt to enter invalid Watchguard ID format\n2. Submit form",
            "Expected Result": "Error displayed: 'Watchguard ID must be 7 alphanumeric characters'. Form submission blocked",
            "Actual Result": "",
            "Status": "",
            "Priority": "Medium",
            "Is Automated": "Yes",
            "Requirement": "REQ2-Validation"
        },
        {
            "TID": "REQ2-UI-006",
            "Scenario": "Watchguard ID Cannot Be Manually Modified",
            "Test Data": "Existing Watchguard ID in record",
            "Test Case Description": "Verify Watchguard ID field is read-only and cannot be modified by users",
            "Pre-Condition": "User has edit permissions on record",
            "Test Steps": "1. Open record detail page\n2. Attempt to modify Watchguard ID field\n3. Try to save changes",
            "Expected Result": "Watchguard ID field is read-only/disabled. No modifications allowed. System ensures immutability",
            "Actual Result": "",
            "Status": "",
            "Priority": "High",
            "Is Automated": "Yes",
            "Requirement": "REQ2-Security"
        },
        {
            "TID": "REQ2-UI-007",
            "Scenario": "Copy Watchguard ID to Clipboard",
            "Test Data": "Watchguard ID: WG7A2B9",
            "Test Case Description": "Verify user can easily copy Watchguard ID to clipboard",
            "Pre-Condition": "Record detail page open",
            "Test Steps": "1. Click 'Copy' button next to Watchguard ID\n2. Paste in text editor to verify",
            "Expected Result": "Watchguard ID copied to clipboard. Toast notification: 'Copied to clipboard'. No extra characters copied",
            "Actual Result": "",
            "Status": "",
            "Priority": "Low",
            "Is Automated": "Yes",
            "Requirement": "REQ2-UX"
        },
        {
            "TID": "REQ2-UI-008",
            "Scenario": "Watchguard ID Displayed in Transaction List",
            "Test Data": "Transaction records with Watchguard IDs",
            "Test Case Description": "Verify Watchguard ID is primary identifier in transaction/order list views",
            "Pre-Condition": "Transaction list view accessible",
            "Test Steps": "1. Navigate to transaction list\n2. Verify column headers\n3. Check if Watchguard ID is displayed (not Shopper ID)",
            "Expected Result": "Watchguard ID shown as primary column. Shopper ID not visible (unless in separate legacy reference column)",
            "Actual Result": "",
            "Status": "",
            "Priority": "High",
            "Is Automated": "Yes",
            "Requirement": "REQ2-AC1"
        },
        {
            "TID": "REQ2-UI-009",
            "Scenario": "Verify Migration Completion (No Shopper ID as Primary)",
            "Test Data": "All migrated records",
            "Test Case Description": "Verify that post-migration, Shopper ID is no longer used as primary identifier in UI",
            "Pre-Condition": "Migration completed for all records",
            "Test Steps": "1. Sample 10 random records\n2. Check primary ID field in each\n3. Verify all show Watchguard ID (not Shopper ID)",
            "Expected Result": "All records use Watchguard ID as primary identifier. No Shopper ID in primary ID field",
            "Actual Result": "",
            "Status": "",
            "Priority": "High",
            "Is Automated": "No",
            "Requirement": "REQ2-Migration"
        },
        {
            "TID": "REQ2-UI-010",
            "Scenario": "Watchguard ID Auto-generation Performance",
            "Test Data": "Bulk create 100 new records",
            "Test Case Description": "Verify Watchguard ID generation is fast and handles bulk operations",
            "Pre-Condition": "Bulk create functionality available",
            "Test Steps": "1. Initiate bulk creation of 100 records\n2. Monitor completion time\n3. Verify all have unique Watchguard IDs",
            "Expected Result": "100 unique Watchguard IDs generated within 5 seconds. All IDs unique and valid. No collisions",
            "Actual Result": "",
            "Status": "",
            "Priority": "Medium",
            "Is Automated": "No",
            "Requirement": "REQ2-Performance"
        }
    ]

def get_req2_api_tests():
    """API test cases for Requirement 2: Watchguard ID"""
    return [
        {
            "TID": "REQ2-API-001",
            "Endpoint": "/api/v1/customers",
            "HTTP Method": "POST",
            "Request Body": "{\"name\": \"John Doe\", \"email\": \"john@example.com\"}",
            "Request Headers": "Authorization: Bearer <token>, Content-Type: application/json",
            "Expected Status": "201",
            "Expected Response": "{\"watchguardId\": \"WG7A2B9\", \"name\": \"John Doe\", \"shopperId\": null}",
            "Test Data": "Valid customer creation data",
            "Actual Status": "",
            "Actual Response": "",
            "Status": "",
            "Priority": "High",
            "Requirement": "REQ2-API-Create"
        },
        {
            "TID": "REQ2-API-002",
            "Endpoint": "/api/v1/customers/{watchguardId}",
            "HTTP Method": "GET",
            "Request Body": "N/A",
            "Request Headers": "Authorization: Bearer <token>",
            "Expected Status": "200",
            "Expected Response": "{\"watchguardId\": \"WG7A2B9\", \"name\": \"...\", \"email\": \"...\", \"legacyReference\": {\"originalShopperId\": \"SHP-000123456\"}}",
            "Test Data": "Valid Watchguard ID",
            "Actual Status": "",
            "Actual Response": "",
            "Status": "",
            "Priority": "High",
            "Requirement": "REQ2-API-Retrieve"
        },
        {
            "TID": "REQ2-API-003",
            "Endpoint": "/api/v1/transactions",
            "HTTP Method": "POST",
            "Request Body": "{\"watchguardId\": \"WG7A2B9\", \"amount\": 100.00, \"currency\": \"USD\"}",
            "Request Headers": "Authorization: Bearer <token>, Content-Type: application/json",
            "Expected Status": "201",
            "Expected Response": "{\"transactionId\": \"...\", \"watchguardId\": \"WG7A2B9\", \"mongoDoc\": {\"watchguardId\": \"WG7A2B9\", \"originalShopperId\": \"SHP-000123456\"}}",
            "Test Data": "Valid Watchguard ID, transaction amount",
            "Actual Status": "",
            "Actual Response": "",
            "Status": "",
            "Priority": "High",
            "Requirement": "REQ2-API-Transaction"
        },
        {
            "TID": "REQ2-API-004",
            "Endpoint": "/api/v1/customers/validate-watchguard-id",
            "HTTP Method": "POST",
            "Request Body": "{\"watchguardId\": \"WG7A2B9\"}",
            "Request Headers": "Authorization: Bearer <token>, Content-Type: application/json",
            "Expected Status": "200",
            "Expected Response": "{\"valid\": true, \"format\": \"7-alphanumeric\", \"exists\": true}",
            "Test Data": "Valid Watchguard ID",
            "Actual Status": "",
            "Actual Response": "",
            "Status": "",
            "Priority": "Medium",
            "Requirement": "REQ2-API-Validation"
        },
        {
            "TID": "REQ2-API-005",
            "Endpoint": "/api/v1/customers/validate-watchguard-id",
            "HTTP Method": "POST",
            "Request Body": "{\"watchguardId\": \"INVALID\"}",
            "Request Headers": "Authorization: Bearer <token>, Content-Type: application/json",
            "Expected Status": "400",
            "Expected Response": "{\"valid\": false, \"error\": \"Watchguard ID must be 7 alphanumeric characters\"}",
            "Test Data": "Invalid format (less than 7 chars)",
            "Actual Status": "",
            "Actual Response": "",
            "Status": "",
            "Priority": "High",
            "Requirement": "REQ2-API-Format-Validation"
        },
        {
            "TID": "REQ2-API-006",
            "Endpoint": "/api/v1/customers/search",
            "HTTP Method": "GET",
            "Request Body": "N/A",
            "Request Headers": "Authorization: Bearer <token>",
            "Expected Status": "200",
            "Expected Response": "{\"results\": [{\"watchguardId\": \"WG7A2B9\", ...}], \"totalCount\": 1}",
            "Test Data": "Query parameter: watchguardId=WG7A2B9",
            "Actual Status": "",
            "Actual Response": "",
            "Status": "",
            "Priority": "Medium",
            "Requirement": "REQ2-API-Search"
        },
        {
            "TID": "REQ2-API-007",
            "Endpoint": "/api/v1/watchguard-ids/check-uniqueness",
            "HTTP Method": "POST",
            "Request Body": "{\"watchguardIds\": [\"WG7A2B9\", \"WG7A2B9\"]}",
            "Request Headers": "Authorization: Bearer <token>, Content-Type: application/json",
            "Expected Status": "200",
            "Expected Response": "{\"unique\": false, \"duplicates\": [\"WG7A2B9\"]}",
            "Test Data": "List with duplicate Watchguard IDs",
            "Actual Status": "",
            "Actual Response": "",
            "Status": "",
            "Priority": "High",
            "Requirement": "REQ2-API-Uniqueness"
        },
        {
            "TID": "REQ2-API-008",
            "Endpoint": "/api/v1/migration/shopper-to-watchguard/validate",
            "HTTP Method": "POST",
            "Request Body": "{\"shopperId\": \"SHP-000123456\"}",
            "Request Headers": "Authorization: Bearer <token>, Content-Type: application/json",
            "Expected Status": "200",
            "Expected Response": "{\"migrationValid\": true, \"watchguardId\": \"WG7A2B9\", \"originalShopperIdStored\": true}",
            "Test Data": "Valid migrated Shopper ID",
            "Actual Status": "",
            "Actual Response": "",
            "Status": "",
            "Priority": "High",
            "Requirement": "REQ2-API-Migration-Validation"
        },
        {
            "TID": "REQ2-API-009",
            "Endpoint": "/api/v1/customers/{watchguardId}",
            "HTTP Method": "PUT",
            "Request Body": "{\"watchguardId\": \"WG-MODIFY\"}",
            "Request Headers": "Authorization: Bearer <token>, Content-Type: application/json",
            "Expected Status": "400",
            "Expected Response": "{\"error\": \"Watchguard ID cannot be modified\"}",
            "Test Data": "Attempting to modify existing Watchguard ID",
            "Actual Status": "",
            "Actual Response": "",
            "Status": "",
            "Priority": "High",
            "Requirement": "REQ2-API-Immutability"
        },
        {
            "TID": "REQ2-API-010",
            "Endpoint": "/api/v1/watchguard-ids/generate-batch",
            "HTTP Method": "POST",
            "Request Body": "{\"count\": 100}",
            "Request Headers": "Authorization: Bearer <token>, Content-Type: application/json",
            "Expected Status": "200",
            "Expected Response": "{\"generated\": 100, \"watchguardIds\": [...], \"allUnique\": true, \"generationTime\": \"250ms\"}",
            "Test Data": "Batch generation of 100 IDs",
            "Actual Status": "",
            "Actual Response": "",
            "Status": "",
            "Priority": "Medium",
            "Requirement": "REQ2-API-Batch-Generation"
        }
    ]

def get_integration_tests():
    """Integration and smoke tests"""
    return [
        {
            "TID": "INT-001",
            "Test Type": "Integration",
            "Scenario": "End-to-End Profile ID Linking Workflow",
            "Pre-Condition": "New Watchguard ID created; Multiple Profile IDs exist",
            "Test Steps": "1. Create new customer (generates Watchguard ID)\n2. Link first Profile ID via UI\n3. Link second Profile ID via API\n4. Verify both linked in UI\n5. Verify data consistency in database",
            "Expected Result": "Complete workflow succeeds. UI and API changes sync immediately. Database reflects all links",
            "Priority": "High",
            "Status": "",
            "Requirement": "REQ1-Integration"
        },
        {
            "TID": "INT-002",
            "Test Type": "Integration",
            "Scenario": "Migration Validation - All Data Intact",
            "Pre-Condition": "Legacy Loyalty ID data exists; Migration completed",
            "Test Steps": "1. Run pre-migration data snapshot\n2. Execute migration\n3. Run post-migration validation\n4. Compare record counts and data integrity\n5. Verify no orphaned records",
            "Expected Result": "Zero data loss. All records migrated. All links preserved. Data integrity validated",
            "Priority": "High",
            "Status": "",
            "Requirement": "REQ1-Integration-Migration"
        },
        {
            "TID": "INT-003",
            "Test Type": "Integration",
            "Scenario": "Watchguard ID Propagation Across Systems",
            "Pre-Condition": "Customer created with Watchguard ID; Multiple backend systems exist",
            "Test Steps": "1. Create customer (Watchguard ID generated)\n2. Verify ID in transaction system\n3. Verify ID in reporting system\n4. Verify ID in CRM\n5. Check MongoDB transaction document",
            "Expected Result": "Watchguard ID consistently used across all systems. Original Shopper ID stored only in MongoDB reference field",
            "Priority": "High",
            "Status": "",
            "Requirement": "REQ2-Integration-Propagation"
        },
        {
            "TID": "INT-004",
            "Test Type": "Smoke Test",
            "Scenario": "Basic CRUD Operations - Profile ID",
            "Pre-Condition": "System initialized",
            "Test Steps": "1. Create Profile ID\n2. Read Profile ID\n3. Link to Watchguard ID\n4. Unlink Profile ID\n5. Delete Profile ID",
            "Expected Result": "All CRUD operations succeed without errors",
            "Priority": "High",
            "Status": "",
            "Requirement": "REQ1-Smoke"
        },
        {
            "TID": "INT-005",
            "Test Type": "Smoke Test",
            "Scenario": "Basic CRUD Operations - Watchguard ID",
            "Pre-Condition": "System initialized",
            "Test Steps": "1. Create customer (Watchguard ID auto-generated)\n2. Retrieve customer by Watchguard ID\n3. Update customer data\n4. Search by Watchguard ID\n5. Verify data integrity",
            "Expected Result": "All CRUD operations succeed. Watchguard ID immutable. Search returns correct results",
            "Priority": "High",
            "Status": "",
            "Requirement": "REQ2-Smoke"
        },
        {
            "TID": "INT-006",
            "Test Type": "Integration",
            "Scenario": "Backward Compatibility - Legacy Shopper ID Reference",
            "Pre-Condition": "Migration completed; Old systems may still reference Shopper ID",
            "Test Steps": "1. Query transaction by Watchguard ID\n2. Verify MongoDB document contains originalShopperId\n3. Test API that returns legacy reference\n4. Verify no breaking changes in response",
            "Expected Result": "Watchguard ID used as primary identifier. Legacy Shopper ID accessible via reference field. No breaking changes",
            "Priority": "Medium",
            "Status": "",
            "Requirement": "REQ2-Integration-Compatibility"
        },
        {
            "TID": "INT-007",
            "Test Type": "Integration",
            "Scenario": "Performance - Bulk Profile ID Linking",
            "Pre-Condition": "1000 Profile IDs available; 100 Watchguard IDs ready",
            "Test Steps": "1. Initiate bulk link operation (10 Profile IDs per Watchguard ID)\n2. Monitor performance metrics\n3. Verify completion time\n4. Validate all links created\n5. Check database performance impact",
            "Expected Result": "Bulk operation completes within SLA (< 5 seconds). Zero data loss. Database performance acceptable",
            "Priority": "Medium",
            "Status": "",
            "Requirement": "REQ1-Integration-Performance"
        },
        {
            "TID": "INT-008",
            "Test Type": "Integration",
            "Scenario": "API and UI Consistency Check",
            "Pre-Condition": "Both UI and API endpoints operational",
            "Test Steps": "1. Create Profile ID link via API\n2. Immediately query via UI\n3. Create Watchguard ID via UI\n4. Query via API\n5. Compare results",
            "Expected Result": "Data consistency within 1 second. No sync delays. Both interfaces show identical data",
            "Priority": "Medium",
            "Status": "",
            "Requirement": "REQ1-REQ2-Consistency"
        }
    ]

if __name__ == "__main__":
    create_test_cases_workbook()
